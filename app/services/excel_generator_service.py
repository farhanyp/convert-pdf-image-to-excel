import os
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import List, Dict, Any

class ExcelGeneratorService:
    """Service for generating Excel files from schedule data."""
    
    def __init__(self, logger=None):
        """Initialize the Excel generator service with a logger."""
        self.logger = logger or logging.getLogger(__name__)
    
    def create_standard_excel(self, employees: List[Dict[str, Any]], file_path: str) -> None:
        """
        Create a standard Excel file with employee schedules.
        
        Args:
            employees: List of employee dictionaries with schedule information
            file_path: Path where the Excel file should be saved
        """
        try:
            self.logger.info(f"Creating standard Excel file for {len(employees)} employees")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employee Schedules"
            
            # Set up headers
            headers = ['ID', 'Name', 'Position', 'Department', 'Monday', 'Tuesday', 
                      'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            
            # Style for headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Populate data
            for row_idx, employee in enumerate(employees, 2):
                ws.cell(row=row_idx, column=1).value = employee.get('id')
                ws.cell(row=row_idx, column=2).value = employee.get('name')
                ws.cell(row=row_idx, column=3).value = employee.get('position')
                ws.cell(row=row_idx, column=4).value = employee.get('department')
                
                # Add schedule by day
                schedule = employee.get('schedule', {})
                days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                
                for day_idx, day in enumerate(days, 5):
                    ws.cell(row=row_idx, column=day_idx).value = schedule.get(day, 'Off')
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = max(max_length + 2, 12)  # Min width of 12
                ws.column_dimensions[column].width = adjusted_width
            
            # Save the workbook
            wb.save(file_path)
            self.logger.info(f"Standard Excel file created successfully at {file_path}")
            
        except Exception as e:
            self.logger.error(f"Error creating standard Excel file: {str(e)}")
            raise
    
    def create_printable_excel(self, employees: List[Dict[str, Any]], file_path: str) -> None:
        """
        Create a printable Excel file (A5 format) with employee schedules.
        
        Args:
            employees: List of employee dictionaries with schedule information
            file_path: Path where the Excel file should be saved
        """
        try:
            self.logger.info(f"Creating printable Excel file for {len(employees)} employees")
            
            wb = openpyxl.Workbook()
            
            # Create a worksheet for each employee
            for employee in employees:
                # Create a sheet named after the employee (max 31 chars for sheet name)
                sheet_name = f"{employee.get('name', 'Employee')} ({employee.get('id')})"
                sheet_name = sheet_name[:31]  # Excel sheet name length limit
                
                # Check if sheet name already exists to avoid duplicates
                if sheet_name in wb.sheetnames:
                    sheet_name = f"{sheet_name}_{employee.get('id')}"[:31]
                    
                ws = wb.create_sheet(title=sheet_name)
                
                # Set up A5 page formatting
                ws.page_setup.paperSize = 11  # A5 paper size
                ws.page_setup.orientation = 'portrait'
                
                # Add employee header information
                ws.merge_cells('A1:G1')
                header_cell = ws.cell(row=1, column=1)
                header_cell.value = f"Schedule for: {employee.get('name', 'Employee')}"
                header_cell.font = Font(size=14, bold=True)
                header_cell.alignment = Alignment(horizontal='center')
                
                # Add position and department
                ws.merge_cells('A2:G2')
                info_cell = ws.cell(row=2, column=1)
                info_cell.value = f"Position: {employee.get('position', 'N/A')} | Department: {employee.get('department', 'N/A')}"
                info_cell.alignment = Alignment(horizontal='center')
                
                # Add days of the week header
                ws.merge_cells('A4:G4')
                ws.cell(row=4, column=1).value = "Weekly Schedule"
                ws.cell(row=4, column=1).font = Font(bold=True)
                ws.cell(row=4, column=1).alignment = Alignment(horizontal='center')
                
                # Add each day's schedule
                schedule = employee.get('schedule', {})
                days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                
                for day_idx, day in enumerate(days, 5):
                    ws.merge_cells(f'A{day_idx}:C{day_idx}')
                    day_cell = ws.cell(row=day_idx, column=1)
                    day_cell.value = day
                    day_cell.font = Font(bold=True)
                    
                    ws.merge_cells(f'D{day_idx}:G{day_idx}')
                    schedule_cell = ws.cell(row=day_idx, column=4)
                    schedule_cell.value = schedule.get(day, 'Off')
                
                # Set column widths
                for col in range(1, 8):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10
            
            # Remove the default sheet if there are other sheets
            if len(wb.sheetnames) > 1:
                del wb['Sheet']
            
            # Save the workbook
            wb.save(file_path)
            self.logger.info(f"Printable Excel file created successfully at {file_path}")
            
        except Exception as e:
            self.logger.error(f"Error creating printable Excel file: {str(e)}")
            raise
    
    def format_bytes(self, bytes_value: int, precision: int = 2) -> str:
        """
        Format file size in bytes to human readable format.
        
        Args:
            bytes_value: Size in bytes
            precision: Number of decimal places
            
        Returns:
            Formatted string (e.g., "2.5 MB")
        """
        if bytes_value == 0:
            return "0 B"
            
        units = ['B', 'KB', 'MB', 'GB', 'TB']
        
        i = 0
        while bytes_value >= 1024 and i < len(units) - 1:
            bytes_value /= 1024
            i += 1
            
        return f"{bytes_value:.{precision}f} {units[i]}"